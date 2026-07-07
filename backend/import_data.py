"""
Import de données ESG depuis un fichier CSV ou Excel (100% local).

Format attendu : deux colonnes « Champ ; Valeur » (une ligne par indicateur),
tolérant aux libellés français, aux clés internes et aux variantes de casse /
accents / séparateurs. Un modèle téléchargeable est fourni.
"""
import csv
import io
import unicodedata

# (section, clé interne, type, [libellés/synonymes])
FIELD_SPECS = [
    # ── Entreprise ──────────────────────────────────────────────────────
    ("company", "name", "str", ["nom", "nom de l'entreprise", "entreprise", "raison sociale", "company", "name"]),
    ("company", "sector", "str", ["secteur", "secteur d'activite", "industrie", "sector"]),
    ("company", "country", "str", ["pays", "country"]),
    ("company", "revenue_eur", "num", ["chiffre d'affaires", "ca", "chiffre d affaires", "revenue", "revenue_eur", "turnover"]),
    ("company", "reporting_year", "int", ["annee de reporting", "annee", "exercice", "reporting_year", "year"]),
    ("company", "target_year", "int", ["horizon des objectifs", "annee cible", "target_year", "horizon"]),
    ("company", "presenter_name", "str", ["nom du presentateur", "presentateur", "presenter", "presenter_name"]),
    ("company", "presenter_title", "str", ["fonction du presentateur", "fonction", "presenter_title", "titre"]),
    # ── Environnement ───────────────────────────────────────────────────
    ("environmental", "co2_emissions_tonnes", "num", ["emissions co2 totales", "co2", "co2 total", "emissions co2", "co2_emissions_tonnes", "ges"]),
    ("environmental", "energy_consumption_mwh", "num", ["consommation energie", "consommation d'energie", "energie", "mwh", "energy_consumption_mwh"]),
    ("environmental", "renewable_energy_percent", "num", ["energie renouvelable", "part energie renouvelable", "renouvelable", "renewable_energy_percent", "renewable"]),
    ("environmental", "water_consumption_m3", "num", ["consommation eau", "consommation d'eau", "eau", "eau prelevee", "water_consumption_m3"]),
    ("environmental", "waste_generated_tonnes", "num", ["dechets generes", "dechets", "waste_generated_tonnes"]),
    ("environmental", "waste_recycled_percent", "num", ["taux de recyclage", "recyclage", "waste_recycled_percent"]),
    ("environmental", "biodiversity_initiatives", "int", ["initiatives biodiversite", "biodiversite", "biodiversity_initiatives"]),
    ("environmental", "scope1_emissions", "num", ["scope 1", "scope1", "emissions directes", "scope1_emissions"]),
    ("environmental", "scope2_emissions", "num", ["scope 2", "scope2", "energie indirecte", "scope2_emissions"]),
    ("environmental", "scope3_emissions", "num", ["scope 3", "scope3", "emissions indirectes", "scope3_emissions"]),
    # ── Social ──────────────────────────────────────────────────────────
    ("social", "total_employees", "int", ["effectif total", "effectif", "employes", "collaborateurs", "total_employees"]),
    ("social", "female_employees_percent", "num", ["femmes dans l'effectif", "part de femmes", "femmes", "female_employees_percent"]),
    ("social", "employee_turnover_percent", "num", ["turnover", "taux de turnover", "rotation", "employee_turnover_percent"]),
    ("social", "training_hours_per_employee", "num", ["formation", "heures de formation", "formation par employe", "training_hours_per_employee"]),
    ("social", "work_accidents", "int", ["accidents de travail", "accidents", "work_accidents"]),
    ("social", "accident_frequency_rate", "num", ["taux de frequence", "taux frequence accidents", "tf", "accident_frequency_rate"]),
    ("social", "community_investment_eur", "num", ["investissement communaute", "investissement communautaire", "community_investment_eur"]),
    ("social", "local_suppliers_percent", "num", ["fournisseurs locaux", "part fournisseurs locaux", "local_suppliers_percent"]),
    ("social", "customer_satisfaction_score", "num", ["satisfaction client", "satisfaction", "customer_satisfaction_score"]),
    ("social", "disabled_employees_percent", "num", ["salaries handicapes", "handicap", "travailleurs handicapes", "disabled_employees_percent"]),
    # ── Gouvernance ─────────────────────────────────────────────────────
    ("governance", "board_members", "int", ["membres du ca", "membres du conseil", "conseil d'administration", "board_members"]),
    ("governance", "female_board_percent", "num", ["femmes au ca", "femmes au conseil", "female_board_percent"]),
    ("governance", "independent_board_percent", "num", ["administrateurs independants", "independants", "independent_board_percent"]),
    ("governance", "ethics_violations", "int", ["violations ethiques", "manquements ethiques", "ethique", "ethics_violations"]),
    ("governance", "corruption_cases", "int", ["cas de corruption", "corruption", "corruption_cases"]),
    ("governance", "data_breaches", "int", ["violations de donnees", "fuites de donnees", "cybersecurite", "data_breaches"]),
    ("governance", "csr_budget_eur", "num", ["budget rse", "budget csr", "csr_budget_eur"]),
    ("governance", "esg_audit_conducted", "bool", ["audit esg", "audit esg conduit", "audit esg independant", "esg_audit_conducted"]),
    ("governance", "sustainability_committee", "bool", ["comite de durabilite", "comite rse", "comite durable", "sustainability_committee"]),
    # ── Taxonomie ───────────────────────────────────────────────────────
    ("taxonomy", "turnover_aligned_percent", "num", ["ca aligne taxonomie", "ca aligne", "turnover aligne", "turnover_aligned_percent"]),
    ("taxonomy", "capex_aligned_percent", "num", ["capex aligne taxonomie", "capex aligne", "capex_aligned_percent"]),
    ("taxonomy", "opex_aligned_percent", "num", ["opex aligne taxonomie", "opex aligne", "opex_aligned_percent"]),
]


def _norm(s: str) -> str:
    """Normalise : minuscule, sans accents, espaces/ponctuation compactés."""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    for ch in ":%()·":
        s = s.replace(ch, " ")
    return " ".join(s.split())


# index normalisé synonyme -> (section, key, type)
_LOOKUP = {}
for section, key, typ, labels in FIELD_SPECS:
    for lab in labels + [key]:
        _LOOKUP[_norm(lab)] = (section, key, typ)


def _to_number(v):
    s = str(v).strip().replace(" ", "").replace(" ", "")
    # gère « 1 234,56 », « 1,234.56 », « 85000000 € », « 42% »
    s = s.replace("€", "").replace("%", "").replace("t", "").strip()
    if "," in s and "." in s:
        s = s.replace(",", "") if s.rfind(",") < s.rfind(".") else s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


_TRUE = {"oui", "yes", "true", "vrai", "1", "o", "y", "x"}
_FALSE = {"non", "no", "false", "faux", "0", "n"}


def _coerce(typ, raw):
    if raw is None or str(raw).strip() == "":
        return None
    if typ == "str":
        return str(raw).strip()
    if typ == "bool":
        n = _norm(raw)
        if n in _TRUE:
            return True
        if n in _FALSE:
            return False
        return None
    num = _to_number(raw)
    if num is None:
        return None
    return int(round(num)) if typ == "int" else num


def _pairs_from_rows(rows):
    """Extrait des couples (champ, valeur) depuis une liste de lignes.

    Accepte le format « clé;valeur » (2 colonnes) et le format large
    (ligne d'en-têtes + 1 ligne de valeurs)."""
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return []
    # Format large : beaucoup de colonnes, >= 2 lignes, plusieurs en-têtes reconnus
    if len(rows) >= 2:
        header = rows[0]
        recognized = sum(1 for c in header if _norm(c) in _LOOKUP)
        if len(header) > 2 and recognized >= 2:
            return list(zip(header, rows[1]))
    # Format clé/valeur : 2 premières colonnes de chaque ligne
    return [(r[0], r[1] if len(r) > 1 else "") for r in rows]


def build_form(pairs):
    """Construit les sections du formulaire depuis des couples (champ, valeur)."""
    sections = {"company": {}, "environmental": {}, "social": {},
                "governance": {}, "taxonomy": {}}
    matched, unmatched = [], []
    _HEADERS = {"champ", "field", "indicateur", "cle", "key"}
    for field, value in pairs:
        if field is None or str(field).strip() == "":
            continue
        if _norm(field) in _HEADERS and _norm(value) in {"valeur", "value", "exemple", ""}:
            continue  # ligne d'en-tête
        spec = _LOOKUP.get(_norm(field))
        if not spec:
            if str(value).strip():
                unmatched.append(str(field).strip())
            continue
        section, key, typ = spec
        coerced = _coerce(typ, value)
        if coerced is not None:
            sections[section][key] = coerced
            matched.append(key)
    sections = {k: v for k, v in sections.items() if v}
    return {"sections": sections, "matched": matched, "unmatched": unmatched}


def parse_csv(data: bytes):
    text = data.decode("utf-8-sig", errors="replace")
    # détection du séparateur (; ou , ou tab)
    sample = text[:2000]
    delim = ";" if sample.count(";") >= sample.count(",") else ","
    if "\t" in sample and sample.count("\t") > sample.count(delim):
        delim = "\t"
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    return _pairs_from_rows([row for row in reader])


def parse_xlsx(data: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else c for c in row])
    wb.close()
    return _pairs_from_rows(rows)


def parse_upload(filename: str, data: bytes):
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx(data)
    if name.endswith(".xls"):
        raise ValueError("Le format .xls (ancien Excel) n'est pas supporté — enregistrez en .xlsx ou .csv")
    return parse_csv(data)


def template_csv() -> str:
    """Modèle CSV clé/valeur avec exemples, prêt à remplir."""
    examples = {
        "name": "Acme Industries SA", "sector": "Industrie", "country": "France",
        "revenue_eur": "85000000", "reporting_year": "2024", "target_year": "2030",
        "presenter_name": "Marie Dupont", "presenter_title": "Directrice RSE",
        "co2_emissions_tonnes": "8840", "energy_consumption_mwh": "21500",
        "renewable_energy_percent": "42", "water_consumption_m3": "67000",
        "waste_generated_tonnes": "320", "waste_recycled_percent": "71",
        "biodiversity_initiatives": "4", "scope1_emissions": "3200",
        "scope2_emissions": "2100", "scope3_emissions": "3540",
        "total_employees": "480", "female_employees_percent": "44",
        "employee_turnover_percent": "9", "training_hours_per_employee": "22",
        "work_accidents": "3", "accident_frequency_rate": "2.8",
        "community_investment_eur": "120000", "local_suppliers_percent": "60",
        "customer_satisfaction_score": "8.1", "disabled_employees_percent": "4.2",
        "board_members": "9", "female_board_percent": "44",
        "independent_board_percent": "55", "ethics_violations": "0",
        "corruption_cases": "0", "data_breaches": "0", "csr_budget_eur": "250000",
        "esg_audit_conducted": "Oui", "sustainability_committee": "Oui",
        "turnover_aligned_percent": "38", "capex_aligned_percent": "52",
        "opex_aligned_percent": "29",
    }
    out = io.StringIO()
    w = csv.writer(out, delimiter=";")
    w.writerow(["Champ", "Valeur"])
    for section, key, typ, labels in FIELD_SPECS:
        w.writerow([labels[0].capitalize(), examples.get(key, "")])
    return out.getvalue()
